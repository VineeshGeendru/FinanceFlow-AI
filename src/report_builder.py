"""
report_builder.py — Build the three-tab Excel variance report.

WHAT IT DOES (in finance terms):
  Produces the deliverable that goes to the CFO — a formatted Excel workbook
  with three tabs:

    Tab 1 — Executive Summary
      The "front page." P&L summary table (Revenue → COGS → Gross Profit →
      Opex → Operating Income) with budget, actual, variance, and the AI
      executive summary bullets below it.

    Tab 2 — YTD Variance Analysis
      One row per line item, all closed months summed. Red highlight on
      material unfavorable variances, green on material favorable ones.
      Commentary column showing the AI-generated (or template) explanation.

    Tab 3 — Monthly Detail
      Every row of the monthly variance table with filters on all columns.
      The analyst's working tab — use it to drill into any month or department.

WHY LIVE EXCEL FORMULAS (not hardcoded values):
  The Variance $ and Variance % columns use Excel formulas (=Actual-Budget,
  =Variance/Budget) instead of the Python-computed numbers. This means if a
  finance team member corrects a number in the Actual column, the variance
  updates automatically — the report stays accurate without re-running the tool.
  Hardcoded values would break the moment anyone edits a cell.

INTERVIEW ANGLE:
  "How does your report handle corrections after it's delivered?"
  Answer: variance columns are live formulas, not static values. Editing
  Budget or Actual in any cell automatically recalculates variance dollar
  and percent. The report behaves like a real finance model.
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from src.config import OUTPUT_DIR, REPORT_FILENAME


# ── Color palette ─────────────────────────────────────────────────────────────
DARK_BLUE   = "1F3864"   # header background
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F2F2F2"   # alternating row
GREEN_FILL  = "C6EFCE"   # favorable material variance
RED_FILL    = "FFB3B3"   # unfavorable material variance
YELLOW_FILL = "FFEB9C"   # subtotal rows (Gross Profit, Op Income)
BORDER_GRAY = "BFBFBF"

# ── Number formats ────────────────────────────────────────────────────────────
FMT_DOLLARS  = '$#,##0'
FMT_VARIANCE = '+$#,##0;-$#,##0;$-'   # shows + on positive, - on negative
FMT_PCT      = '+0.0%;-0.0%;0.0%'
FMT_PCT_PLAIN= '0.0%'


# ── Main entry point ──────────────────────────────────────────────────────────

def build_report(results: dict, commentary: dict,
                 output_path: str = None) -> str:
    """
    Build the three-tab Excel report.  Returns the path to the saved file.

    results   — dict from run_variance_engine()
    commentary — dict from generate_commentary()
    output_path — optional override; defaults to output/REPORT_FILENAME
    """
    if output_path is None:
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        output_path = os.path.join(OUTPUT_DIR, REPORT_FILENAME)

    wb = Workbook()

    _build_exec_summary(wb, results, commentary)
    _build_ytd_tab(wb, results, commentary)
    _build_monthly_tab(wb, results)

    # Remove the default blank sheet openpyxl creates
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(output_path)
    print(f"[REPORT] Saved: {output_path}")
    return output_path


# ── Tab 1: Executive Summary ──────────────────────────────────────────────────

def _build_exec_summary(wb: Workbook, results: dict, commentary: dict):
    ws = wb.create_sheet("Executive Summary")
    ws.sheet_view.showGridLines = False

    pl   = results["pl_summary"]
    data = results["monthly"]
    bullets = commentary["executive_summary"]

    # ── Header ────────────────────────────────────────────────────────────────
    closed_months = sorted(data["Month"].unique())
    period = f"{closed_months[0]} to {closed_months[-1]}"

    ws.merge_cells("A1:F1")
    ws["A1"] = "FinanceFlow-AI  |  Variance Analysis Report"
    ws["A1"].font      = Font(bold=True, size=16, color=WHITE)
    ws["A1"].fill      = PatternFill("solid", fgColor=DARK_BLUE)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center",
                                   indent=1)
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:F2")
    ws["A2"] = f"YTD Period: {period}  |  Closed months: {len(closed_months)}"
    ws["A2"].font      = Font(italic=True, size=10, color="666666")
    ws["A2"].alignment = Alignment(horizontal="left", indent=1)
    ws.row_dimensions[2].height = 16

    # ── P&L table header ──────────────────────────────────────────────────────
    row = 4
    headers = ["", "Budget", "Actual", "Variance $", "Variance %", ""]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font      = Font(bold=True, color=WHITE)
        cell.fill      = PatternFill("solid", fgColor=DARK_BLUE)
        cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[row].height = 18

    # ── P&L rows ──────────────────────────────────────────────────────────────
    subtotal_lines = {"Gross Profit", "Operating Income"}
    separator_before = {"Gross Profit", "Opex"}

    for _, pl_row in pl.iterrows():
        line = pl_row["Line"]
        if line in separator_before:
            row += 1  # blank separator row

        row += 1

        # Determine fill
        if line in subtotal_lines:
            fill = PatternFill("solid", fgColor=YELLOW_FILL)
            bold = True
        else:
            fill = PatternFill("solid", fgColor=LIGHT_GRAY) if row % 2 == 0 \
                   else PatternFill("solid", fgColor=WHITE)
            bold = False

        # Variance color on the variance cell
        net_impact = pl_row["Net Impact $"]
        var_color = GREEN_FILL if net_impact > 0 else \
                    (RED_FILL  if net_impact < 0 else WHITE)

        # Col A — line name
        c = ws.cell(row=row, column=1, value=line)
        c.font      = Font(bold=bold)
        c.fill      = fill
        c.alignment = Alignment(indent=1 if line not in subtotal_lines else 0)

        # Col B — Budget
        c = ws.cell(row=row, column=2, value=pl_row["Budget"])
        c.number_format = FMT_DOLLARS
        c.font  = Font(bold=bold)
        c.fill  = fill
        c.alignment = Alignment(horizontal="right")

        # Col C — Actual
        c = ws.cell(row=row, column=3, value=pl_row["Actual"])
        c.number_format = FMT_DOLLARS
        c.font  = Font(bold=bold)
        c.fill  = fill
        c.alignment = Alignment(horizontal="right")

        # Col D — Variance $ (live formula: =C{row}-B{row})
        c = ws.cell(row=row, column=4,
                    value=f"=C{row}-B{row}")
        c.number_format = FMT_VARIANCE
        c.font  = Font(bold=bold)
        c.fill  = PatternFill("solid", fgColor=var_color)
        c.alignment = Alignment(horizontal="right")

        # Col E — Variance % (live formula: =IF(B{row}<>0,D{row}/B{row},0))
        c = ws.cell(row=row, column=5,
                    value=f"=IF(B{row}<>0,D{row}/B{row},0)")
        c.number_format = FMT_PCT
        c.font  = Font(bold=bold)
        c.fill  = PatternFill("solid", fgColor=var_color)
        c.alignment = Alignment(horizontal="right")

        # Col F — Favorable/Unfavorable label
        c = ws.cell(row=row, column=6, value=pl_row["F/U"])
        c.font  = Font(bold=bold,
                       color="375623" if net_impact > 0 else
                             ("9C0006" if net_impact < 0 else "000000"))
        c.fill  = fill
        c.alignment = Alignment(horizontal="center")

    # ── Executive summary commentary ──────────────────────────────────────────
    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "Management Commentary"
    ws[f"A{row}"].font      = Font(bold=True, size=12, color=WHITE)
    ws[f"A{row}"].fill      = PatternFill("solid", fgColor=DARK_BLUE)
    ws[f"A{row}"].alignment = Alignment(horizontal="left", indent=1)
    ws.row_dimensions[row].height = 20

    for bullet in bullets:
        row += 1
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = f"•  {bullet}"
        ws[f"A{row}"].alignment = Alignment(
            wrap_text=True, vertical="top", indent=1
        )
        ws.row_dimensions[row].height = 42

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 16


# ── Tab 2: YTD Variance Analysis ──────────────────────────────────────────────

def _build_ytd_tab(wb: Workbook, results: dict, commentary: dict):
    ws = wb.create_sheet("YTD Variance Analysis")
    ws.sheet_view.showGridLines = False

    ytd = results["ytd"].copy()
    line_commentary = commentary["line_commentary"]

    # Add commentary column to YTD DataFrame
    # We match on Department + Line Item (YTD doesn't have a single Month key,
    # so we use the most recent month's commentary for each line item).
    def get_ytd_commentary(dept, line_item):
        # Find any commentary key matching this dept + line item
        matches = {k: v for k, v in line_commentary.items()
                   if k[1] == dept and k[2] == line_item}
        if not matches:
            return ""
        # Return the latest month's commentary
        return matches[max(matches.keys(), key=lambda k: k[0])]

    # ── Header row ────────────────────────────────────────────────────────────
    headers = [
        "Department", "Line Item", "Category",
        "YTD Budget", "YTD Actual",
        "Variance $", "Variance %",
        "Fav / Unfav", "Material", "Commentary"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = Font(bold=True, color=WHITE)
        cell.fill      = PatternFill("solid", fgColor=DARK_BLUE)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[1].height = 20

    # ── Data rows ─────────────────────────────────────────────────────────────
    for i, (_, row) in enumerate(ytd.iterrows(), start=2):
        is_material = row["YTD_Material"]
        is_fav      = row["YTD_Favorable"]

        if is_material:
            fill_color = GREEN_FILL if is_fav else RED_FILL
        else:
            fill_color = LIGHT_GRAY if i % 2 == 0 else WHITE

        row_fill = PatternFill("solid", fgColor=fill_color)
        budget_col = 4  # Col D

        values = [
            row["Department"],
            row["Line Item"],
            row["Category"],
            row["YTD_Budget"],    # Col D — Budget (value)
            row["YTD_Actual"],    # Col E — Actual (value)
            None,                 # Col F — Variance $ (formula)
            None,                 # Col G — Variance % (formula)
            row["YTD_FU"],
            "Yes" if is_material else "No",
            get_ytd_commentary(row["Department"], row["Line Item"]),
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.fill      = row_fill
            cell.alignment = Alignment(horizontal="left", wrap_text=(col == 10))

        # Live formulas for variance columns
        ws.cell(row=i, column=6,
                value=f"=E{i}-D{i}").number_format = FMT_VARIANCE
        ws.cell(row=i, column=6).fill = row_fill
        ws.cell(row=i, column=6).alignment = Alignment(horizontal="right")

        ws.cell(row=i, column=7,
                value=f"=IF(D{i}<>0,F{i}/D{i},0)").number_format = FMT_PCT
        ws.cell(row=i, column=7).fill = row_fill
        ws.cell(row=i, column=7).alignment = Alignment(horizontal="right")

        # Dollar format on Budget and Actual
        ws.cell(row=i, column=4).number_format = FMT_DOLLARS
        ws.cell(row=i, column=4).alignment = Alignment(horizontal="right")
        ws.cell(row=i, column=5).number_format = FMT_DOLLARS
        ws.cell(row=i, column=5).alignment = Alignment(horizontal="right")

        ws.row_dimensions[i].height = 40 if len(str(values[9])) > 60 else 18

    # ── Auto-filter ───────────────────────────────────────────────────────────
    ws.auto_filter.ref = f"A1:J{len(ytd) + 1}"

    # ── Column widths ─────────────────────────────────────────────────────────
    widths = [22, 24, 12, 14, 14, 14, 12, 14, 10, 55]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # ── Legend ────────────────────────────────────────────────────────────────
    legend_row = len(ytd) + 3
    ws.cell(row=legend_row, column=1, value="Legend:")
    ws.cell(row=legend_row, column=1).font = Font(bold=True)
    items = [
        (GREEN_FILL, "Material Favorable"),
        (RED_FILL,   "Material Unfavorable"),
        (LIGHT_GRAY, "Non-material"),
    ]
    for offset, (color, label) in enumerate(items, 1):
        c = ws.cell(row=legend_row, column=1 + offset, value=label)
        c.fill = PatternFill("solid", fgColor=color)
        c.alignment = Alignment(horizontal="center")


# ── Tab 3: Monthly Detail ─────────────────────────────────────────────────────

def _build_monthly_tab(wb: Workbook, results: dict):
    ws = wb.create_sheet("Monthly Detail")
    ws.sheet_view.showGridLines = False

    monthly = results["monthly"].copy()

    headers = [
        "Month", "Department", "Line Item", "Category",
        "Budget", "Actual",
        "Variance $", "Variance %",
        "Net Impact $", "Fav / Unfav", "Material"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = Font(bold=True, color=WHITE)
        cell.fill      = PatternFill("solid", fgColor=DARK_BLUE)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[1].height = 20

    col_map = {
        "Month": 1, "Department": 2, "Line Item": 3, "Category": 4,
        "Amount_budget": 5, "Amount_actual": 6,
        "Net Impact $": 9, "F/U": 10,
    }

    for i, (_, row) in enumerate(monthly.iterrows(), start=2):
        is_material = row["Material"]
        is_fav      = row["Favorable"]

        if is_material:
            fill_color = GREEN_FILL if is_fav else RED_FILL
        else:
            fill_color = LIGHT_GRAY if i % 2 == 0 else WHITE
        row_fill = PatternFill("solid", fgColor=fill_color)

        # Static value columns
        for src_col, dest_col in col_map.items():
            val = row[src_col]
            cell = ws.cell(row=i, column=dest_col, value=val)
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal="left")

        # Budget and Actual formatting
        ws.cell(row=i, column=5).number_format  = FMT_DOLLARS
        ws.cell(row=i, column=5).alignment       = Alignment(horizontal="right")
        ws.cell(row=i, column=6).number_format  = FMT_DOLLARS
        ws.cell(row=i, column=6).alignment       = Alignment(horizontal="right")

        # Live formulas — Variance $ and Variance %
        ws.cell(row=i, column=7,
                value=f"=F{i}-E{i}").number_format = FMT_VARIANCE
        ws.cell(row=i, column=7).fill      = row_fill
        ws.cell(row=i, column=7).alignment = Alignment(horizontal="right")

        ws.cell(row=i, column=8,
                value=f"=IF(E{i}<>0,G{i}/E{i},0)").number_format = FMT_PCT
        ws.cell(row=i, column=8).fill      = row_fill
        ws.cell(row=i, column=8).alignment = Alignment(horizontal="right")

        # Net Impact formatting
        ws.cell(row=i, column=9).number_format  = FMT_VARIANCE
        ws.cell(row=i, column=9).alignment       = Alignment(horizontal="right")
        ws.cell(row=i, column=9).fill            = row_fill

        # Material column
        mat_val = "Yes" if is_material else "No"
        c = ws.cell(row=i, column=11, value=mat_val)
        c.fill = row_fill
        c.alignment = Alignment(horizontal="center")

        ws.row_dimensions[i].height = 15

    # ── Auto-filter and freeze panes ──────────────────────────────────────────
    ws.auto_filter.ref = f"A1:K{len(monthly) + 1}"
    ws.freeze_panes    = "A2"   # keep header visible when scrolling

    # ── Column widths ─────────────────────────────────────────────────────────
    widths = [12, 22, 24, 12, 14, 14, 14, 12, 14, 14, 10]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
